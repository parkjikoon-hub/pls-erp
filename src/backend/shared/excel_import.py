"""
공통 Excel 임포트 엔진
어떤 엑셀 파일이든 컬럼 매핑을 통해 PLS ERP 데이터로 변환합니다.
이카운트, 수기 엑셀 등 출처에 상관없이 사용 가능합니다.

사용법:
    1. parse_excel_headers() → 엑셀 컬럼 목록 추출
    2. auto_match_columns() → 자동 매핑 제안
    3. import_excel_data() → 매핑 정보로 데이터 변환 + 검증
"""
import io
import re
from typing import Optional
from openpyxl import load_workbook
from pydantic import BaseModel, Field


class ColumnMapping(BaseModel):
    """하나의 컬럼 매핑 정보"""
    excel_column: str = Field(description="엑셀 파일의 컬럼명")
    target_field: Optional[str] = Field(None, description="PLS ERP 필드명 (None이면 무시)")


class FieldDefinition(BaseModel):
    """PLS ERP 필드 정의 — 자동 매핑 + 검증에 사용"""
    field_name: str = Field(description="DB 필드명")
    display_name: str = Field(description="화면 표시명")
    required: bool = Field(default=False, description="필수 여부")
    aliases: list[str] = Field(default_factory=list, description="자동 매핑용 별칭 (이카운트 컬럼명 등)")
    transform: Optional[str] = Field(None, description="변환 규칙 (business_no 등)")


class ImportResult(BaseModel):
    """임포트 결과"""
    success_count: int = 0
    fail_count: int = 0
    errors: list[dict] = Field(default_factory=list)
    total_rows: int = 0


# ── 각 모듈별 필드 정의 (이카운트 컬럼명을 aliases에 포함) ──

CUSTOMER_FIELDS: list[FieldDefinition] = [
    FieldDefinition(field_name="code", display_name="거래처코드", required=True,
                    aliases=["거래처코드", "코드", "거래처 코드", "Code", "거래처CD"]),
    FieldDefinition(field_name="name", display_name="거래처명", required=True,
                    aliases=["거래처명", "거래처", "상호", "업체명", "회사명", "Name"]),
    FieldDefinition(field_name="business_no", display_name="사업자등록번호",
                    aliases=["사업자번호", "사업자등록번호", "사업자 등록번호", "사업자No", "BusinessNo"], transform="business_no"),
    FieldDefinition(field_name="ceo_name", display_name="대표자명",
                    aliases=["대표자", "대표자명", "대표", "CEO"]),
    FieldDefinition(field_name="business_type", display_name="업태",
                    aliases=["업태", "업종", "BusinessType"]),
    FieldDefinition(field_name="business_item", display_name="종목",
                    aliases=["종목", "업종(종목)", "Item", "업종2"]),
    FieldDefinition(field_name="address", display_name="주소",
                    aliases=["주소", "주소1", "사업장주소", "Address"]),
    FieldDefinition(field_name="phone", display_name="전화번호",
                    aliases=["전화", "전화번호", "TEL", "전화1", "Phone"]),
    FieldDefinition(field_name="email", display_name="이메일",
                    aliases=["이메일", "메일", "E-Mail", "Email"]),
    FieldDefinition(field_name="fax", display_name="팩스",
                    aliases=["팩스", "FAX", "팩스번호"]),
    FieldDefinition(field_name="contact_person", display_name="담당자",
                    aliases=["담당자", "담당자명", "담당", "Contact"]),
    FieldDefinition(field_name="customer_type", display_name="거래처유형",
                    aliases=["유형", "거래유형", "거래처유형", "구분", "매출매입구분"]),
    FieldDefinition(field_name="credit_limit", display_name="신용한도",
                    aliases=["신용한도", "여신한도", "한도"]),
    FieldDefinition(field_name="payment_terms", display_name="결제조건(일)",
                    aliases=["결제조건", "결제일", "지급조건"]),
    FieldDefinition(field_name="bank_name", display_name="은행명",
                    aliases=["은행", "은행명", "거래은행"]),
    FieldDefinition(field_name="bank_account", display_name="계좌번호",
                    aliases=["계좌번호", "계좌", "통장번호"]),
    FieldDefinition(field_name="bank_account_name", display_name="예금주",
                    aliases=["예금주", "예금주명"]),
]

PRODUCT_FIELDS: list[FieldDefinition] = [
    FieldDefinition(field_name="code", display_name="품목코드", required=True,
                    aliases=["품목코드", "품목 코드", "코드", "품번", "Code", "품목CD"]),
    FieldDefinition(field_name="name", display_name="품목명", required=True,
                    aliases=["품목명", "품명", "상품명", "제품명", "Name"]),
    FieldDefinition(field_name="product_type", display_name="품목유형",
                    aliases=["유형", "품목유형", "구분", "자재구분"]),
    FieldDefinition(field_name="unit", display_name="단위",
                    aliases=["단위", "기본단위", "Unit", "규격단위"]),
    FieldDefinition(field_name="standard_price", display_name="기준단가",
                    aliases=["단가", "기준단가", "판매단가", "표준단가", "Price"]),
    FieldDefinition(field_name="cost_price", display_name="원가",
                    aliases=["원가", "매입단가", "구매단가", "Cost"]),
    FieldDefinition(field_name="safety_stock", display_name="안전재고",
                    aliases=["안전재고", "최소재고", "적정재고"]),
    FieldDefinition(field_name="tax_rate", display_name="부가세율(%)",
                    aliases=["세율", "부가세율", "VAT", "부가세"]),
]

# 모듈별 필드 정의 레지스트리 — 새 모듈 추가 시 여기에 등록
FIELD_REGISTRY: dict[str, list[FieldDefinition]] = {
    "customers": CUSTOMER_FIELDS,
    "products": PRODUCT_FIELDS,
    # Phase 2 이후 추가 예정:
    # "inventory": INVENTORY_FIELDS,
    # "sales_orders": SALES_ORDER_FIELDS,
    # "production_orders": PRODUCTION_FIELDS,
    # "shipments": SHIPMENT_FIELDS,
}


def parse_excel_headers(file_bytes: bytes, sheet_name: Optional[str] = None) -> dict:
    """
    엑셀 파일을 읽어 시트 목록과 컬럼 헤더를 추출합니다.

    반환 형식:
    {
        "sheets": ["Sheet1", "Sheet2"],
        "selected_sheet": "Sheet1",
        "headers": ["거래처코드", "거래처명", "사업자번호", ...],
        "preview_rows": [["C001", "삼성전자", "123-45-67890", ...], ...],
        "total_rows": 150,
    }
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheets = wb.sheetnames

    # 시트 선택 (지정 없으면 첫 번째 시트)
    target_sheet = sheet_name if sheet_name and sheet_name in sheets else sheets[0]
    ws = wb[target_sheet]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return {"sheets": sheets, "selected_sheet": target_sheet, "headers": [], "preview_rows": [], "total_rows": 0}

    # 첫 번째 행을 헤더로 사용
    headers = [str(cell).strip() if cell is not None else f"열{i+1}" for i, cell in enumerate(rows[0])]

    # 미리보기 (최대 5행)
    preview_rows = []
    for row in rows[1:6]:
        preview_rows.append([str(cell) if cell is not None else "" for cell in row])

    return {
        "sheets": sheets,
        "selected_sheet": target_sheet,
        "headers": headers,
        "preview_rows": preview_rows,
        "total_rows": len(rows) - 1,  # 헤더 제외
    }


def auto_match_columns(excel_headers: list[str], module: str) -> list[ColumnMapping]:
    """
    엑셀 컬럼명과 PLS ERP 필드를 자동 매칭합니다.
    이카운트 컬럼명, 유사 이름, 정확한 이름 모두 매칭 시도합니다.
    """
    fields = FIELD_REGISTRY.get(module, [])
    mappings = []

    for header in excel_headers:
        header_clean = header.strip()
        matched_field = None

        for field in fields:
            # 1. 정확한 aliases 매칭
            if header_clean in field.aliases:
                matched_field = field.field_name
                break

            # 2. 대소문자 무시 매칭
            if header_clean.lower() in [a.lower() for a in field.aliases]:
                matched_field = field.field_name
                break

            # 3. 필드 표시명과 매칭
            if header_clean == field.display_name:
                matched_field = field.field_name
                break

            # 4. 부분 매칭 (해더에 별칭이 포함되어 있을 때)
            for alias in field.aliases:
                if alias in header_clean or header_clean in alias:
                    matched_field = field.field_name
                    break
            if matched_field:
                break

        mappings.append(ColumnMapping(excel_column=header, target_field=matched_field))

    return mappings


def import_excel_data(
    file_bytes: bytes,
    module: str,
    column_mappings: list[ColumnMapping],
    sheet_name: Optional[str] = None,
) -> tuple[list[dict], list[dict]]:
    """
    매핑 정보를 기반으로 엑셀 데이터를 PLS ERP 형식으로 변환합니다.

    반환: (valid_rows, error_rows)
      - valid_rows: 정상 변환된 데이터 목록
      - error_rows: 에러가 있는 행 (행번호 + 사유 포함)
    """
    fields = FIELD_REGISTRY.get(module, [])
    field_map = {f.field_name: f for f in fields}

    # 매핑 딕셔너리 생성: 엑셀 컬럼 인덱스 → 타깃 필드
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    target_sheet = sheet_name if sheet_name and sheet_name in wb.sheetnames else wb.sheetnames[0]
    ws = wb[target_sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return [], []

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]

    # 매핑 인덱스 구축
    col_to_field: dict[int, str] = {}
    for mapping in column_mappings:
        if mapping.target_field and mapping.excel_column in headers:
            idx = headers.index(mapping.excel_column)
            col_to_field[idx] = mapping.target_field

    valid_rows = []
    error_rows = []

    for row_num, row in enumerate(rows[1:], start=2):  # 2행부터 (헤더가 1행)
        record = {}
        row_errors = []

        for col_idx, target_field in col_to_field.items():
            cell_value = row[col_idx] if col_idx < len(row) else None
            value = str(cell_value).strip() if cell_value is not None else ""

            # 빈 값 처리
            if value == "" or value == "None":
                record[target_field] = None
                continue

            # 변환 규칙 적용
            field_def = field_map.get(target_field)
            if field_def and field_def.transform:
                value = _apply_transform(value, field_def.transform, row_errors, row_num, target_field)

            # 숫자 필드 변환
            if target_field in ("credit_limit", "standard_price", "cost_price", "tax_rate"):
                try:
                    value = float(re.sub(r"[^\d.\-]", "", str(value)))
                except ValueError:
                    row_errors.append(f"{target_field}: 숫자가 아닙니다 ({value})")
                    value = 0
            elif target_field in ("payment_terms", "safety_stock"):
                try:
                    value = int(float(re.sub(r"[^\d.\-]", "", str(value))))
                except ValueError:
                    row_errors.append(f"{target_field}: 숫자가 아닙니다 ({value})")
                    value = 0
            elif target_field == "customer_type":
                value = _normalize_customer_type(value)
            elif target_field == "product_type":
                value = _normalize_product_type(value)

            record[target_field] = value

        # 필수 필드 검증
        for field in fields:
            if field.required and not record.get(field.field_name):
                row_errors.append(f"필수 항목 '{field.display_name}'이(가) 비어있습니다")

        if row_errors:
            error_rows.append({
                "row": row_num,
                "data": record,
                "errors": row_errors,
            })
        else:
            valid_rows.append(record)

    return valid_rows, error_rows


def _apply_transform(value: str, transform: str, errors: list, row_num: int, field: str) -> str:
    """필드별 변환 규칙 적용"""
    if transform == "business_no":
        # 사업자등록번호: 숫자만 추출 → 000-00-00000 형식
        digits = re.sub(r"[^0-9]", "", value)
        if len(digits) == 10:
            return f"{digits[:3]}-{digits[3:5]}-{digits[5:]}"
        elif digits:
            errors.append(f"사업자등록번호 형식 오류 (10자리 필요, {len(digits)}자리 입력)")
            return value
    return value


def _normalize_customer_type(value: str) -> str:
    """이카운트 거래처 유형 → PLS ERP 유형으로 변환"""
    v = value.strip().lower()
    mapping = {
        "매출": "customer", "매출처": "customer", "판매": "customer", "고객": "customer",
        "매입": "supplier", "매입처": "supplier", "구매": "supplier", "공급": "supplier", "공급처": "supplier",
        "겸용": "both", "매출매입": "both", "매출/매입": "both", "전체": "both",
        "customer": "customer", "supplier": "supplier", "both": "both",
    }
    return mapping.get(v, "both")


def _normalize_product_type(value: str) -> str:
    """이카운트 품목 유형 → PLS ERP 유형으로 변환"""
    v = value.strip().lower()
    mapping = {
        "제품": "product", "상품": "product", "완제품": "product",
        "자재": "material", "원자재": "material", "부자재": "material", "원재료": "material",
        "반제품": "semi", "반완제품": "semi",
        "product": "product", "material": "material", "semi": "semi",
    }
    return mapping.get(v, "product")
