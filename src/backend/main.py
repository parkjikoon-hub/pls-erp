"""
PLS ERP — FastAPI 메인 애플리케이션
모든 모듈 라우터를 등록하고 미들웨어를 설정합니다.
프론트엔드(React 빌드)도 같은 서버에서 제공합니다.
"""
from pathlib import Path

from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse
from contextlib import asynccontextmanager

from .config import settings
from .database import engine, AsyncSessionLocal


@asynccontextmanager
async def lifespan(app: FastAPI):
    """앱 시작/종료 시 수행 작업"""
    import logging
    _log = logging.getLogger(__name__)

    # 시작 1: 누락된 DB 테이블/컬럼 자동 생성
    try:
        from .database import Base
        # 모든 모듈의 모델을 import (테이블 메타데이터 등록)
        from .modules.m1_system import models as _m1  # noqa: F401
        from .modules.m4_finance import models as _m4  # noqa: F401
        from .modules.m3_hr import models as _m3  # noqa: F401
        from .modules.m2_sales import models as _m2  # noqa: F401
        from .modules.m5_production import models as _m5  # noqa: F401
        from .modules.m6_groupware import models as _m6  # noqa: F401
        from .modules.m7_notifications import models as _m7  # noqa: F401
        # 누락된 테이블 자동 생성 (기존 테이블은 건드리지 않음)
        async with engine.begin() as conn:
            await conn.run_sync(Base.metadata.create_all)
        _log.info("DB 테이블 동기화 완료 (누락 테이블 자동 생성)")
    except Exception as e:
        _log.warning(f"DB 테이블 동기화 실패 (무시): {e}")

    # 시작 2: 회사 정보 시드 데이터 (최초 1회)
    try:
        from .modules.m1_system.models import CompanyInfo
        from sqlalchemy import select as sa_select
        async with AsyncSessionLocal() as db:
            existing = await db.execute(sa_select(CompanyInfo).limit(1))
            if existing.scalar_one_or_none() is None:
                db.add(CompanyInfo(
                    company_name="(주)피엘에스",
                    business_no="704-88-01943",
                    corp_no="195411-0039707",
                    ceo_name="위봉길, 강현철",
                    address="경상남도 진주시 사봉면 산업단지로43번길 39",
                    business_type="제조업",
                    business_item="소방기구, 배관",
                ))
                await db.commit()
                _log.info("회사 정보 시드 데이터 생성 완료")
    except Exception as e:
        _log.warning(f"회사 정보 시드 데이터 생성 실패 (무시): {e}")

    # 시작 3: M5 창고 시드 데이터 (실패해도 앱은 계속 실행)
    try:
        from .modules.m5_production.seed import seed_warehouses
        async with AsyncSessionLocal() as db:
            await seed_warehouses(db)
            await db.commit()
    except Exception as e:
        _log.warning(f"시드 데이터 초기화 실패 (무시): {e}")

    yield
    # 종료: DB 연결 정리
    await engine.dispose()


app = FastAPI(
    title=settings.APP_NAME,
    description="PLS ERP — AI 통합 ERP REST API",
    version=settings.APP_VERSION,
    docs_url="/api/docs",
    redoc_url="/api/redoc",
    lifespan=lifespan,
)

# CORS 설정 (프론트엔드에서 API 호출 허용, 여러 도메인 지원)
_origins = [o.strip() for o in settings.FRONTEND_URL.split(",") if o.strip()]
app.add_middleware(
    CORSMiddleware,
    allow_origins=_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# 헬스체크 엔드포인트 (서버 동작 확인용)
@app.get("/api/health", tags=["시스템"])
async def health_check():
    """서버 상태 확인"""
    return {
        "status": "healthy",
        "app": settings.APP_NAME,
        "version": settings.APP_VERSION,
    }


@app.get("/api/diag/openpyxl", tags=["시스템"])
async def diag_openpyxl():
    """openpyxl 진단 — Excel 생성 기능 점검용 (임시)"""
    import sys
    result = {"python": sys.version}
    try:
        import openpyxl
        result["openpyxl_version"] = openpyxl.__version__
    except Exception as e:
        result["openpyxl_error"] = str(e)
        return result
    try:
        import io as _io
        from openpyxl.styles import Font
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="test")
        ws.cell(row=1, column=1).font = Font(name="맑은 고딕", size=10)
        buf = _io.BytesIO()
        wb.save(buf)
        result["excel_bytes"] = len(buf.getvalue())
        result["status"] = "OK"
    except Exception as e:
        result["excel_error"] = str(e)
    return result


# ── 모듈 라우터 등록 (개발 순서에 따라 순차 추가) ──

# Phase 1: M1 시스템 아키텍처 & MDM
from .auth.router import router as auth_router
app.include_router(auth_router, prefix="/api/v1/auth", tags=["인증"])

from .modules.m1_system.router import router as m1_router
app.include_router(m1_router, prefix="/api/v1/system", tags=["M1-시스템"])

from .audit.router import router as audit_router
app.include_router(audit_router, prefix="/api/v1/audit", tags=["감사로그"])

# Phase 2: M4 재무/회계
from .modules.m4_finance.routers import router as m4_router
app.include_router(m4_router, prefix="/api/v1/finance", tags=["M4-재무회계"])

# Phase 3: M3 인사/급여
from .modules.m3_hr.routers import router as m3_router
app.include_router(m3_router, prefix="/api/v1/hr", tags=["M3-인사급여"])

# Phase 4: M2 영업/수주
from .modules.m2_sales.routers import router as m2_router
app.include_router(m2_router, prefix="/api/v1/sales", tags=["M2-영업수주"])

# Phase 5: M5 생산/SCM
from .modules.m5_production.routers import router as m5_router
app.include_router(m5_router, prefix="/api/v1/production", tags=["M5-생산SCM"])

# Phase 6: M6 그룹웨어
from .modules.m6_groupware.routers import router as m6_router
app.include_router(m6_router, prefix="/api/v1/groupware", tags=["M6-그룹웨어"])

# Phase 7: M7 알림센터
from .modules.m7_notifications.routers import router as m7_router
app.include_router(m7_router, prefix="/api/v1/notifications", tags=["M7-알림센터"])

# 대시보드 요약 API
from .dashboard.router import router as dashboard_router
app.include_router(dashboard_router, prefix="/api/v1/dashboard", tags=["대시보드"])


# ── 프론트엔드 정적 파일 서빙 (React SPA) ──
# React 빌드 결과물(dist/)을 FastAPI에서 직접 제공합니다.
# 모든 API 라우터보다 아래에 있어야 API 경로가 우선 처리됩니다.
_frontend_dist = Path(__file__).resolve().parent.parent / "frontend" / "dist"

if _frontend_dist.is_dir():
    # CSS, JS, 이미지 등 정적 파일 제공 (/assets/...)
    app.mount("/assets", StaticFiles(directory=_frontend_dist / "assets"), name="assets")

    # SPA 폴백: API가 아닌 모든 경로 → index.html 반환
    @app.get("/{full_path:path}", include_in_schema=False)
    async def serve_spa(full_path: str):
        """React SPA 라우팅 지원 — 존재하는 파일이면 반환, 아니면 index.html"""
        file_path = _frontend_dist / full_path
        if file_path.is_file():
            return FileResponse(file_path)
        return FileResponse(_frontend_dist / "index.html")
