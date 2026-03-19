"""
M2 영업/수주 — 수주 서비스 (비즈니스 로직)
"""
import uuid
from datetime import date
from typing import Optional

from sqlalchemy import select, func, or_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from fastapi import HTTPException, status

from ..models import SalesOrder, SalesOrderLine, Quotation, QuotationLine
from ..schemas.orders import SalesOrderCreate, SalesOrderUpdate
from ....audit.service import log_action
from ...m1_system.models import Customer


async def _generate_order_no(db: AsyncSession, order_date: date) -> str:
    """수주번호 자동 생성 (SO-YYYYMM-NNNN)"""
    prefix = f"SO-{order_date.strftime('%Y%m')}-"
    result = await db.execute(
        select(func.count())
        .select_from(SalesOrder)
        .where(SalesOrder.order_no.like(f"{prefix}%"))
    )
    seq = (result.scalar() or 0) + 1
    return f"{prefix}{seq:04d}"


async def list_orders(
    db: AsyncSession,
    customer_id: Optional[str] = None,
    status_filter: Optional[str] = None,
    search: Optional[str] = None,
    page: int = 1,
    size: int = 20,
):
    """수주 목록 조회"""
    query = select(SalesOrder).where(SalesOrder.is_active.is_(True))

    if customer_id:
        query = query.where(SalesOrder.customer_id == uuid.UUID(customer_id))
    if status_filter:
        query = query.where(SalesOrder.status == status_filter)
    if search:
        sf = f"%{search}%"
        query = query.where(or_(
            SalesOrder.order_no.ilike(sf),
            SalesOrder.notes.ilike(sf),
        ))

    count_q = select(func.count()).select_from(query.subquery())
    total = (await db.execute(count_q)).scalar() or 0

    query = (
        query
        .options(selectinload(SalesOrder.customer), selectinload(SalesOrder.sales_rep))
        .order_by(SalesOrder.created_at.desc())
        .offset((page - 1) * size)
        .limit(size)
    )
    items = (await db.execute(query)).scalars().all()

    return {
        "items": [
            {
                "id": str(o.id),
                "order_no": o.order_no,
                "order_date": o.order_date,
                "delivery_date": o.delivery_date,
                "customer_name": o.customer.name if o.customer else None,
                "sales_rep_name": o.sales_rep.name if o.sales_rep else None,
                "total_amount": float(o.total_amount or 0),
                "tax_amount": float(o.tax_amount or 0),
                "grand_total": float(o.grand_total or 0),
                "status": o.status,
                "progress_pct": o.progress_pct,
                "created_at": o.created_at,
            }
            for o in items
        ],
        "total": total,
        "page": page,
        "size": size,
        "total_pages": (total + size - 1) // size if total > 0 else 1,
    }


async def get_order(db: AsyncSession, order_id: uuid.UUID):
    """수주 상세 조회"""
    result = await db.execute(
        select(SalesOrder)
        .options(
            selectinload(SalesOrder.lines).selectinload(SalesOrderLine.product),
            selectinload(SalesOrder.customer),
            selectinload(SalesOrder.sales_rep),
            selectinload(SalesOrder.quotation),
        )
        .where(SalesOrder.id == order_id, SalesOrder.is_active.is_(True))
    )
    o = result.scalar_one_or_none()
    if not o:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "수주를 찾을 수 없습니다")

    return {
        "id": str(o.id),
        "order_no": o.order_no,
        "order_date": o.order_date,
        "delivery_date": o.delivery_date,
        "customer_id": str(o.customer_id),
        "customer_name": o.customer.name if o.customer else None,
        "quotation_id": str(o.quotation_id) if o.quotation_id else None,
        "quotation_no": o.quotation.quote_no if o.quotation else None,
        "sales_rep_id": str(o.sales_rep_id) if o.sales_rep_id else None,
        "sales_rep_name": o.sales_rep.name if o.sales_rep else None,
        "total_amount": float(o.total_amount or 0),
        "tax_amount": float(o.tax_amount or 0),
        "grand_total": float(o.grand_total or 0),
        "status": o.status,
        "progress_pct": o.progress_pct,
        "notes": o.notes,
        "lines": [
            {
                "id": str(ln.id),
                "line_no": ln.line_no,
                "product_id": str(ln.product_id) if ln.product_id else None,
                "product_name": ln.product_name,
                "specification": ln.specification,
                "quantity": float(ln.quantity),
                "unit_price": float(ln.unit_price),
                "amount": float(ln.amount),
                "tax_amount": float(ln.tax_amount or 0),
                "delivery_date": ln.delivery_date,
                "produced_qty": float(ln.produced_qty or 0),
                "shipped_qty": float(ln.shipped_qty or 0),
                "remark": ln.remark,
            }
            for ln in o.lines
        ],
        "created_at": o.created_at,
    }


async def create_order(
    db: AsyncSession,
    data: SalesOrderCreate,
    current_user,
    ip_address: Optional[str] = None,
):
    """수주 생성 (신규 또는 견적서 기반)"""
    order_no = await _generate_order_no(db, data.order_date)

    order = SalesOrder(
        order_no=order_no,
        order_date=data.order_date,
        delivery_date=data.delivery_date,
        customer_id=uuid.UUID(data.customer_id),
        quotation_id=uuid.UUID(data.quotation_id) if data.quotation_id else None,
        sales_rep_id=uuid.UUID(data.sales_rep_id) if data.sales_rep_id else None,
        notes=data.notes,
        status="confirmed",
        created_by=current_user.id,
    )
    db.add(order)
    await db.flush()

    total_amount = 0.0
    total_tax = 0.0
    for idx, line_data in enumerate(data.lines, 1):
        amount = round(line_data.quantity * line_data.unit_price, 2)
        tax = round(amount * 0.1, 2)
        line = SalesOrderLine(
            order_id=order.id,
            line_no=idx,
            product_id=uuid.UUID(line_data.product_id) if line_data.product_id else None,
            product_name=line_data.product_name,
            specification=line_data.specification,
            quantity=line_data.quantity,
            unit_price=line_data.unit_price,
            amount=amount,
            tax_amount=tax,
            delivery_date=line_data.delivery_date,
            remark=line_data.remark,
        )
        db.add(line)
        total_amount += amount
        total_tax += tax

    order.total_amount = round(total_amount, 2)
    order.tax_amount = round(total_tax, 2)
    order.grand_total = round(total_amount + total_tax, 2)
    await db.flush()

    # 견적서 상태를 accepted로 변경
    if data.quotation_id:
        qt_result = await db.execute(
            select(Quotation).where(Quotation.id == uuid.UUID(data.quotation_id))
        )
        qt = qt_result.scalar_one_or_none()
        if qt and qt.status in ("draft", "sent"):
            qt.status = "accepted"

    await log_action(
        db=db, table_name="sales_orders", record_id=order.id,
        action="INSERT", changed_by=current_user.id,
        new_values={"order_no": order_no, "customer_id": data.customer_id},
        ip_address=ip_address,
    )

    # ── M7 알림: 수주 등록 알림 ──
    from ....shared.notification_helper import notify_order_created
    cust = await db.get(Customer, uuid.UUID(data.customer_id))
    cust_name = cust.name if cust else "알 수 없음"
    await notify_order_created(db, order_no, cust_name, order.id, current_user.id)

    result_data = {"id": str(order.id), "order_no": order_no}

    # ── 자동 작업지시서 생성 (옵션) ──
    if data.auto_create_wo:
        try:
            from ...m5_production.services.work_order_service import create_from_order
            wo_result = await create_from_order(db, order.id, current_user, ip_address)
            result_data["work_orders"] = wo_result.get("work_orders", [])
            result_data["material_shortage"] = wo_result.get("material_shortage", [])
            result_data["has_shortage"] = wo_result.get("has_shortage", False)
        except Exception:
            # 작업지시서 생성 실패해도 수주는 정상 반환
            result_data["wo_error"] = "작업지시서 자동 생성 중 오류가 발생했습니다."

    return result_data


async def create_order_from_quotation(
    db: AsyncSession,
    quotation_id: uuid.UUID,
    order_date: date,
    current_user,
    ip_address: Optional[str] = None,
):
    """견적서 → 수주 전환"""
    result = await db.execute(
        select(Quotation)
        .options(selectinload(Quotation.lines))
        .where(Quotation.id == quotation_id, Quotation.is_active.is_(True))
    )
    qt = result.scalar_one_or_none()
    if not qt:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "견적서를 찾을 수 없습니다")
    if qt.status not in ("draft", "sent", "accepted"):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "rejected 상태의 견적서는 수주로 전환할 수 없습니다")

    order_no = await _generate_order_no(db, order_date)

    order = SalesOrder(
        order_no=order_no,
        order_date=order_date,
        delivery_date=None,
        customer_id=qt.customer_id,
        quotation_id=qt.id,
        sales_rep_id=qt.sales_rep_id,
        notes=f"견적서 {qt.quote_no} 기반 수주",
        status="confirmed",
        total_amount=float(qt.total_amount or 0),
        tax_amount=float(qt.tax_amount or 0),
        grand_total=float(qt.grand_total or 0),
        created_by=current_user.id,
    )
    db.add(order)
    await db.flush()

    # 견적서 라인 → 수주 라인 복사
    for ql in qt.lines:
        line = SalesOrderLine(
            order_id=order.id,
            line_no=ql.line_no,
            product_id=ql.product_id,
            product_name=ql.product_name,
            specification=ql.specification,
            quantity=float(ql.quantity),
            unit_price=float(ql.unit_price),
            amount=float(ql.amount),
            tax_amount=float(ql.tax_amount or 0),
            delivery_date=ql.delivery_date,
        )
        db.add(line)

    # 견적서 상태 → accepted
    qt.status = "accepted"
    await db.flush()

    await log_action(
        db=db, table_name="sales_orders", record_id=order.id,
        action="INSERT", changed_by=current_user.id,
        new_values={"order_no": order_no, "from_quotation": qt.quote_no},
        ip_address=ip_address, memo="견적서 기반 수주 생성",
    )

    # ── M7 알림: 수주 등록 알림 ──
    from ....shared.notification_helper import notify_order_created
    cust = await db.get(Customer, qt.customer_id)
    cust_name = cust.name if cust else "알 수 없음"
    await notify_order_created(db, order_no, cust_name, order.id, current_user.id)

    result_data: dict = {"id": str(order.id), "order_no": order_no}

    # ── 스마트 분기: 재고 체크 → 자동 작업지시서/출하 안내 ──
    try:
        from ...m5_production.services.inventory_service import check_quotation_materials
        lines_for_check = [
            {"product_id": str(ql.product_id), "quantity": float(ql.quantity)}
            for ql in qt.lines if ql.product_id
        ]
        if lines_for_check:
            stock_check = await check_quotation_materials(db, lines_for_check)
            summary = stock_check.get("summary", {})
            result_data["stock_check"] = stock_check

            if summary.get("ready_to_ship"):
                # 완제품 충분 → 출하 준비 가능
                result_data["smart_action"] = "ready_to_ship"
                result_data["smart_message"] = "완제품 재고 충분 — 출하 준비 가능합니다."
            else:
                # 완제품 부족 → 자동 작업지시서 생성
                result_data["smart_action"] = "production_needed"
                try:
                    from ...m5_production.services.work_order_service import create_from_order
                    wo_result = await create_from_order(db, order.id, current_user, ip_address)
                    result_data["work_orders"] = wo_result.get("work_orders", [])
                    result_data["material_shortage"] = wo_result.get("material_shortage", [])
                    result_data["has_shortage"] = wo_result.get("has_shortage", False)

                    if wo_result.get("has_shortage"):
                        result_data["smart_message"] = "생산 필요 + 원자재 구매 필요 — 작업지시서 생성됨, 원자재 구매 알림 발송"
                    else:
                        result_data["smart_message"] = "생산 필요 — 원자재 확보됨, 작업지시서 생성 완료"
                except Exception:
                    result_data["smart_message"] = "생산 필요 — 작업지시서 자동 생성 중 오류 발생"
    except Exception:
        # 재고 체크 실패해도 수주 전환은 정상 완료
        result_data["smart_action"] = "unknown"
        result_data["smart_message"] = "재고 확인 중 오류가 발생했습니다."

    return result_data


async def update_order(
    db: AsyncSession,
    order_id: uuid.UUID,
    data: SalesOrderUpdate,
    current_user,
    ip_address: Optional[str] = None,
):
    """수주 수정 (confirmed 상태만)"""
    result = await db.execute(
        select(SalesOrder)
        .options(selectinload(SalesOrder.lines))
        .where(SalesOrder.id == order_id, SalesOrder.is_active.is_(True))
    )
    o = result.scalar_one_or_none()
    if not o:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "수주를 찾을 수 없습니다")
    if o.status != "confirmed":
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "confirmed 상태에서만 수정할 수 있습니다")

    if data.order_date is not None:
        o.order_date = data.order_date
    if data.delivery_date is not None:
        o.delivery_date = data.delivery_date
    if data.customer_id is not None:
        o.customer_id = uuid.UUID(data.customer_id)
    if data.sales_rep_id is not None:
        o.sales_rep_id = uuid.UUID(data.sales_rep_id) if data.sales_rep_id else None
    if data.notes is not None:
        o.notes = data.notes

    if data.lines is not None:
        for old_line in o.lines:
            await db.delete(old_line)
        await db.flush()

        total_amount = 0.0
        total_tax = 0.0
        for idx, line_data in enumerate(data.lines, 1):
            amount = round(line_data.quantity * line_data.unit_price, 2)
            tax = round(amount * 0.1, 2)
            line = SalesOrderLine(
                order_id=o.id,
                line_no=idx,
                product_id=uuid.UUID(line_data.product_id) if line_data.product_id else None,
                product_name=line_data.product_name,
                specification=line_data.specification,
                quantity=line_data.quantity,
                unit_price=line_data.unit_price,
                amount=amount,
                tax_amount=tax,
                delivery_date=line_data.delivery_date,
                remark=line_data.remark,
            )
            db.add(line)
            total_amount += amount
            total_tax += tax

        o.total_amount = round(total_amount, 2)
        o.tax_amount = round(total_tax, 2)
        o.grand_total = round(total_amount + total_tax, 2)

    await db.flush()

    await log_action(
        db=db, table_name="sales_orders", record_id=o.id,
        action="UPDATE", changed_by=current_user.id,
        new_values={"order_no": o.order_no},
        ip_address=ip_address,
    )

    return {"id": str(o.id), "order_no": o.order_no}


async def update_order_status(
    db: AsyncSession,
    order_id: uuid.UUID,
    new_status: str,
    current_user,
    ip_address: Optional[str] = None,
    memo: Optional[str] = None,
):
    """수주 상태 변경"""
    valid_transitions = {
        "confirmed": ["in_production", "shipped", "completed"],
        "in_production": ["shipped", "completed"],
        "shipped": ["completed", "invoiced"],
        "completed": ["invoiced"],
    }

    result = await db.execute(
        select(SalesOrder).where(
            SalesOrder.id == order_id, SalesOrder.is_active.is_(True)
        )
    )
    o = result.scalar_one_or_none()
    if not o:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "수주를 찾을 수 없습니다")

    allowed = valid_transitions.get(o.status, [])
    if new_status not in allowed:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            f"'{o.status}' → '{new_status}' 상태 전환이 불가합니다. 허용: {allowed}",
        )

    old_status = o.status
    o.status = new_status

    # 상태별 진행률 자동 계산
    progress_map = {
        "confirmed": 0,
        "in_production": 30,
        "shipped": 70,
        "completed": 100,
        "invoiced": 100,
    }
    o.progress_pct = progress_map.get(new_status, o.progress_pct)
    await db.flush()

    await log_action(
        db=db, table_name="sales_orders", record_id=o.id,
        action="UPDATE", changed_by=current_user.id,
        old_values={"status": old_status},
        new_values={"status": new_status},
        ip_address=ip_address,
        memo=memo or f"상태 변경: {old_status} → {new_status}",
    )

    return {"id": str(o.id), "status": new_status, "progress_pct": o.progress_pct}


async def delete_order(
    db: AsyncSession,
    order_id: uuid.UUID,
    current_user,
    ip_address: Optional[str] = None,
):
    """수주 삭제 (confirmed 상태만)"""
    result = await db.execute(
        select(SalesOrder).where(
            SalesOrder.id == order_id, SalesOrder.is_active.is_(True)
        )
    )
    o = result.scalar_one_or_none()
    if not o:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "수주를 찾을 수 없습니다")
    if o.status != "confirmed":
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "confirmed 상태에서만 삭제할 수 있습니다")

    o.is_active = False

    await log_action(
        db=db, table_name="sales_orders", record_id=o.id,
        action="DELETE", changed_by=current_user.id,
        old_values={"order_no": o.order_no},
        ip_address=ip_address,
    )

    return {"message": f"수주 {o.order_no}이(가) 삭제되었습니다"}


async def generate_statement_excel(db: AsyncSession, order_id: uuid.UUID) -> bytes:
    """거래명세서 Excel 양식 생성 (한국 표준 양식)"""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── 데이터 조회 ──
    result = await db.execute(
        select(SalesOrder)
        .options(
            selectinload(SalesOrder.lines).selectinload(SalesOrderLine.product),
            selectinload(SalesOrder.customer),
            selectinload(SalesOrder.sales_rep),
        )
        .where(SalesOrder.id == order_id, SalesOrder.is_active.is_(True))
    )
    o = result.scalar_one_or_none()
    if not o:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "수주를 찾을 수 없습니다")

    # 회사 정보 조회
    from ...m1_system.models import CompanyInfo
    ci_result = await db.execute(select(CompanyInfo).limit(1))
    ci = ci_result.scalar_one_or_none()

    # 거래처 정보
    cust = o.customer

    # ── 워크북 생성 ──
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "거래명세서"

    # 스타일 정의
    thin = Side(style="thin")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font = Font(name="맑은 고딕", size=18, bold=True)
    header_font = Font(name="맑은 고딕", size=10, bold=True)
    normal_font = Font(name="맑은 고딕", size=9)
    label_font = Font(name="맑은 고딕", size=9, bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left_al = Alignment(horizontal="left", vertical="center")
    right_al = Alignment(horizontal="right", vertical="center")
    header_fill = PatternFill(start_color="D9E2F3", fill_type="solid")
    label_fill = PatternFill(start_color="F2F2F2", fill_type="solid")

    # 열 너비 (A~H)
    col_widths = [5, 6, 22, 12, 10, 13, 15, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # ── 제목 ──
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    cell = ws.cell(row=row, column=1, value="거  래  명  세  서")
    cell.font = title_font
    cell.alignment = center
    row += 2

    # ── 공급자 / 공급받는자 정보 (좌우 병렬) ──
    # 공급자 (좌측: col 1~4)
    ci_name = ci.company_name if ci else "(주)피엘에스"
    ci_biz_no = ci.business_no if ci else ""
    ci_ceo = ci.ceo_name if ci else ""
    ci_addr = ci.address if ci else ""
    ci_type = ci.business_type if ci else ""
    ci_item = ci.business_item if ci else ""
    ci_phone = ci.phone if ci else ""
    ci_fax = ci.fax if ci else ""

    # 공급받는자 (우측: col 5~8)
    cu_name = cust.name if cust else ""
    cu_biz_no = cust.business_no if cust else ""
    cu_ceo = cust.ceo_name if cust else ""
    cu_addr = cust.address if cust else ""
    cu_type = cust.business_type if cust else ""
    cu_item = cust.business_item if cust else ""

    # 헤더 행
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c = ws.cell(row=row, column=1, value="공 급 자")
    c.font = header_font; c.alignment = center; c.fill = header_fill; c.border = border_all
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=8)
    c = ws.cell(row=row, column=5, value="공 급 받 는 자")
    c.font = header_font; c.alignment = center; c.fill = header_fill; c.border = border_all
    for col_idx in range(1, 9):
        ws.cell(row=row, column=col_idx).border = border_all
    row += 1

    # 정보 행 (라벨 + 값)
    info_rows = [
        ("상    호", ci_name, "상    호", cu_name),
        ("사업자번호", ci_biz_no, "사업자번호", cu_biz_no),
        ("대 표 자", ci_ceo, "대 표 자", cu_ceo),
        ("주    소", ci_addr, "주    소", cu_addr),
        ("업태/종목", f"{ci_type} / {ci_item}", "업태/종목", f"{cu_type or ''} / {cu_item or ''}"),
        ("TEL/FAX", f"{ci_phone or ''} / {ci_fax or ''}", "", ""),
    ]
    for l_label, l_val, r_label, r_val in info_rows:
        # 좌측 라벨
        c = ws.cell(row=row, column=1, value=l_label)
        c.font = label_font; c.alignment = right_al; c.fill = label_fill; c.border = border_all
        # 좌측 값
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        c = ws.cell(row=row, column=2, value=l_val)
        c.font = normal_font; c.alignment = left_al; c.border = border_all
        for ci2 in range(2, 5):
            ws.cell(row=row, column=ci2).border = border_all
        # 우측 라벨
        c = ws.cell(row=row, column=5, value=r_label)
        c.font = label_font; c.alignment = right_al; c.fill = label_fill; c.border = border_all
        # 우측 값
        ws.merge_cells(start_row=row, start_column=6, end_row=row, end_column=8)
        c = ws.cell(row=row, column=6, value=r_val)
        c.font = normal_font; c.alignment = left_al; c.border = border_all
        for ci2 in range(5, 9):
            ws.cell(row=row, column=ci2).border = border_all
        row += 1

    row += 1

    # ── 거래 정보 ──
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    ws.cell(row=row, column=1, value="수주번호").font = label_font
    ws.cell(row=row, column=1).alignment = right_al
    ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
    ws.cell(row=row, column=3, value=o.order_no).font = normal_font
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
    ws.cell(row=row, column=5, value="거래일자").font = label_font
    ws.cell(row=row, column=5).alignment = right_al
    ws.merge_cells(start_row=row, start_column=7, end_row=row, end_column=8)
    ws.cell(row=row, column=7, value=str(o.order_date)).font = normal_font
    row += 2

    # ── 품목 테이블 헤더 ──
    headers = ["No", "", "품명", "규격", "수량", "단가", "금액", "비고"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = header_font
        cell.alignment = center
        cell.fill = header_fill
        cell.border = border_all
    # No 열 merge
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    row += 1

    # ── 품목 라인 ──
    for ln in o.lines:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.cell(row=row, column=1, value=ln.line_no).font = normal_font
        ws.cell(row=row, column=1).alignment = center
        ws.cell(row=row, column=3, value=ln.product_name).font = normal_font
        ws.cell(row=row, column=4, value=ln.specification or "").font = normal_font
        ws.cell(row=row, column=5, value=float(ln.quantity)).font = normal_font
        ws.cell(row=row, column=5).alignment = right_al
        ws.cell(row=row, column=5).number_format = '#,##0'
        ws.cell(row=row, column=6, value=float(ln.unit_price)).font = normal_font
        ws.cell(row=row, column=6).alignment = right_al
        ws.cell(row=row, column=6).number_format = '#,##0'
        ws.cell(row=row, column=7, value=float(ln.amount)).font = normal_font
        ws.cell(row=row, column=7).alignment = right_al
        ws.cell(row=row, column=7).number_format = '#,##0'
        ws.cell(row=row, column=8, value=ln.remark or "").font = normal_font
        for col_idx in range(1, 9):
            ws.cell(row=row, column=col_idx).border = border_all
        row += 1

    # 빈 줄 채우기 (최소 10줄)
    for _ in range(max(0, 10 - len(o.lines))):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        for col_idx in range(1, 9):
            ws.cell(row=row, column=col_idx).border = border_all
        row += 1

    row += 1

    # ── 합계 영역 ──
    summary = [
        ("공급가액", float(o.total_amount or 0)),
        ("세    액", float(o.tax_amount or 0)),
        ("합    계", float(o.grand_total or 0)),
    ]
    for label, value in summary:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = header_font
        cell.alignment = right_al
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=8)
        val_cell = ws.cell(row=row, column=5, value=value)
        val_cell.font = Font(name="맑은 고딕", size=11, bold=True) if "합" in label else normal_font
        val_cell.alignment = right_al
        val_cell.number_format = '₩#,##0'
        for col_idx in range(1, 9):
            ws.cell(row=row, column=col_idx).border = border_all
        row += 1

    # ── 인수/인도자 ──
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value="인수자:                          (인)").font = normal_font
    ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=8)
    ws.cell(row=row, column=5, value="인도자:                          (인)").font = normal_font

    # ── 인쇄 설정 ──
    ws.print_area = f"A1:H{row + 1}"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    # ── 저장 ──
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
