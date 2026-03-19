"""
M2 영업 — 판매가 관리 서비스
- 거래처별 특별 단가 CRUD
- 엑셀 업로드 (기본가 / 거래처별)
- 가격 조회 (우선순위: 거래처별 > 기본가)
"""
import uuid
import io
from datetime import date
from typing import Optional

from sqlalchemy import select, func, and_, or_, delete
from sqlalchemy.ext.asyncio import AsyncSession
from fastapi import HTTPException, status

from ..models import CustomerPriceList
from ..schemas.price_lists import PriceListCreate, PriceListUpdate
from ...m1_system.models import Product, Customer
from ....audit.service import log_action


# ── 판매가 목록 조회 ──

async def list_price_lists(
    db: AsyncSession,
    customer_id: Optional[str] = None,
    product_id: Optional[str] = None,
    search: Optional[str] = None,
    page: int = 1,
    size: int = 50,
):
    """거래처별 판매가 목록 조회"""
    query = (
        select(CustomerPriceList)
        .join(Product, CustomerPriceList.product_id == Product.id)
        .join(Customer, CustomerPriceList.customer_id == Customer.id)
        .where(CustomerPriceList.is_active.is_(True))
    )

    if customer_id:
        query = query.where(CustomerPriceList.customer_id == uuid.UUID(customer_id))
    if product_id:
        query = query.where(CustomerPriceList.product_id == uuid.UUID(product_id))
    if search:
        sf = f"%{search}%"
        query = query.where(or_(
            Product.name.ilike(sf),
            Product.code.ilike(sf),
            Customer.name.ilike(sf),
        ))

    count_q = select(func.count()).select_from(query.subquery())
    total = (await db.execute(count_q)).scalar() or 0

    from sqlalchemy.orm import selectinload
    query = (
        query
        .options(
            selectinload(CustomerPriceList.customer),
            selectinload(CustomerPriceList.product),
        )
        .order_by(Customer.name, Product.code)
        .offset((page - 1) * size)
        .limit(size)
    )
    items = (await db.execute(query)).scalars().all()

    return {
        "items": [
            {
                "id": str(p.id),
                "customer_id": str(p.customer_id),
                "customer_name": p.customer.name if p.customer else None,
                "product_id": str(p.product_id),
                "product_name": p.product.name if p.product else None,
                "product_code": p.product.code if p.product else None,
                "unit_price": float(p.unit_price),
                "standard_price": float(p.product.standard_price) if p.product else 0,
                "valid_from": p.valid_from.isoformat() if p.valid_from else None,
                "valid_until": p.valid_until.isoformat() if p.valid_until else None,
                "notes": p.notes,
            }
            for p in items
        ],
        "total": total,
        "page": page,
        "size": size,
        "total_pages": (total + size - 1) // size if total > 0 else 1,
    }


# ── 판매가 단건 생성 ──

async def create_price_list(
    db: AsyncSession,
    data: PriceListCreate,
    current_user,
    ip_address: Optional[str] = None,
):
    """거래처별 판매가 등록"""
    pl = CustomerPriceList(
        customer_id=uuid.UUID(data.customer_id),
        product_id=uuid.UUID(data.product_id),
        unit_price=data.unit_price,
        valid_from=data.valid_from,
        valid_until=data.valid_until,
        notes=data.notes,
        created_by=current_user.id,
    )
    db.add(pl)
    await db.flush()

    await log_action(
        db=db, table_name="customer_price_lists", record_id=pl.id,
        action="INSERT", changed_by=current_user.id,
        new_values={"customer_id": data.customer_id, "product_id": data.product_id, "unit_price": data.unit_price},
        ip_address=ip_address,
    )

    return {"id": str(pl.id), "message": "판매가 등록 완료"}


# ── 판매가 수정 ──

async def update_price_list(
    db: AsyncSession,
    price_list_id: uuid.UUID,
    data: PriceListUpdate,
    current_user,
    ip_address: Optional[str] = None,
):
    """거래처별 판매가 수정"""
    pl = await db.get(CustomerPriceList, price_list_id)
    if not pl or not pl.is_active:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "판매가를 찾을 수 없습니다")

    old_values = {"unit_price": float(pl.unit_price)}
    if data.unit_price is not None:
        pl.unit_price = data.unit_price
    if data.valid_from is not None:
        pl.valid_from = data.valid_from
    if data.valid_until is not None:
        pl.valid_until = data.valid_until
    if data.notes is not None:
        pl.notes = data.notes

    await db.flush()

    await log_action(
        db=db, table_name="customer_price_lists", record_id=pl.id,
        action="UPDATE", changed_by=current_user.id,
        old_values=old_values,
        new_values={"unit_price": float(pl.unit_price)},
        ip_address=ip_address,
    )

    return {"id": str(pl.id), "message": "판매가 수정 완료"}


# ── 판매가 삭제 (논리 삭제) ──

async def delete_price_list(
    db: AsyncSession,
    price_list_id: uuid.UUID,
    current_user,
    ip_address: Optional[str] = None,
):
    """거래처별 판매가 삭제"""
    pl = await db.get(CustomerPriceList, price_list_id)
    if not pl or not pl.is_active:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "판매가를 찾을 수 없습니다")

    pl.is_active = False
    await db.flush()

    await log_action(
        db=db, table_name="customer_price_lists", record_id=pl.id,
        action="DELETE", changed_by=current_user.id,
        old_values={"unit_price": float(pl.unit_price)},
        ip_address=ip_address,
    )

    return {"message": "판매가 삭제 완료"}


# ── 가격 조회 (우선순위 적용) ──

async def get_price(
    db: AsyncSession,
    customer_id: uuid.UUID,
    product_id: uuid.UUID,
    ref_date: Optional[date] = None,
) -> dict:
    """
    품목 가격 조회 — 우선순위:
    1순위: customer_price_lists (거래처+품목+유효기간)
    2순위: products.standard_price (기본 판매가)
    3순위: 0
    """
    today = ref_date or date.today()

    # 1순위: 거래처별 특별단가
    query = (
        select(CustomerPriceList)
        .where(
            CustomerPriceList.customer_id == customer_id,
            CustomerPriceList.product_id == product_id,
            CustomerPriceList.is_active.is_(True),
        )
    )
    # 유효기간 필터 (NULL이면 무제한)
    query = query.where(
        or_(
            CustomerPriceList.valid_from.is_(None),
            CustomerPriceList.valid_from <= today,
        ),
        or_(
            CustomerPriceList.valid_until.is_(None),
            CustomerPriceList.valid_until >= today,
        ),
    )
    query = query.order_by(CustomerPriceList.created_at.desc()).limit(1)
    result = await db.execute(query)
    pl = result.scalar_one_or_none()

    if pl:
        return {
            "unit_price": float(pl.unit_price),
            "source": "customer_special",
        }

    # 2순위: 기본 판매가
    product = await db.get(Product, product_id)
    if product and product.standard_price and float(product.standard_price) > 0:
        return {
            "unit_price": float(product.standard_price),
            "source": "standard",
        }

    # 3순위: 가격 없음
    return {"unit_price": 0, "source": "none"}


# ── 거래처 기준 전체 품목 가격 목록 ──

async def get_customer_prices(
    db: AsyncSession,
    customer_id: uuid.UUID,
    ref_date: Optional[date] = None,
) -> list[dict]:
    """거래처 기준 모든 활성 품목의 가격 조회 (특별단가 > 기본가)"""
    today = ref_date or date.today()

    # 모든 활성 품목 조회
    products_result = await db.execute(
        select(Product).where(Product.is_active.is_(True)).order_by(Product.code)
    )
    products = products_result.scalars().all()

    # 해당 거래처의 유효한 특별단가 일괄 조회
    pl_result = await db.execute(
        select(CustomerPriceList).where(
            CustomerPriceList.customer_id == customer_id,
            CustomerPriceList.is_active.is_(True),
            or_(CustomerPriceList.valid_from.is_(None), CustomerPriceList.valid_from <= today),
            or_(CustomerPriceList.valid_until.is_(None), CustomerPriceList.valid_until >= today),
        )
    )
    special_prices = {pl.product_id: float(pl.unit_price) for pl in pl_result.scalars().all()}

    result = []
    for prod in products:
        if prod.id in special_prices:
            price = special_prices[prod.id]
            source = "customer_special"
        elif prod.standard_price and float(prod.standard_price) > 0:
            price = float(prod.standard_price)
            source = "standard"
        else:
            price = 0
            source = "none"

        result.append({
            "product_id": str(prod.id),
            "product_name": prod.name,
            "product_code": prod.code,
            "unit": prod.unit or "EA",
            "unit_price": price,
            "standard_price": float(prod.standard_price) if prod.standard_price else 0,
            "source": source,
        })

    return result


# ── 엑셀 업로드: 기본 판매가 ──

async def upload_standard_prices(
    db: AsyncSession,
    file_bytes: bytes,
    current_user,
    ip_address: Optional[str] = None,
) -> dict:
    """
    기본 판매가 엑셀 업로드 — products.standard_price 일괄 업데이트
    엑셀 형식: [품목코드, 품목명(무시), 판매가]
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active
    if not ws:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "엑셀 시트를 읽을 수 없습니다")

    updated = 0
    errors = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue

        code = str(row[0]).strip()
        try:
            price = float(row[2]) if len(row) > 2 and row[2] else 0
        except (ValueError, TypeError):
            errors.append(f"행 {row_idx}: 가격 형식 오류")
            continue

        # 품목 코드로 조회
        result = await db.execute(
            select(Product).where(Product.code == code, Product.is_active.is_(True))
        )
        product = result.scalar_one_or_none()
        if not product:
            errors.append(f"행 {row_idx}: 품목코드 '{code}' 없음")
            continue

        product.standard_price = price
        updated += 1

    await db.flush()
    wb.close()

    return {
        "message": f"기본 판매가 {updated}건 업데이트 완료",
        "updated": updated,
        "errors": errors,
    }


# ── 엑셀 업로드: 거래처별 특별단가 ──

async def upload_customer_prices(
    db: AsyncSession,
    file_bytes: bytes,
    current_user,
    ip_address: Optional[str] = None,
) -> dict:
    """
    거래처별 특별단가 엑셀 업로드 — customer_price_lists upsert
    엑셀 형식: [거래처코드, 품목코드, 판매가, 유효시작일(선택), 유효종료일(선택)]
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active
    if not ws:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "엑셀 시트를 읽을 수 없습니다")

    created = 0
    updated = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            continue

        customer_code = str(row[0]).strip()
        product_code = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        try:
            price = float(row[2]) if len(row) > 2 and row[2] else 0
        except (ValueError, TypeError):
            errors.append(f"행 {row_idx}: 가격 형식 오류")
            continue

        # 유효기간 파싱 (선택)
        valid_from = None
        valid_until = None
        if len(row) > 3 and row[3]:
            try:
                valid_from = row[3] if isinstance(row[3], date) else date.fromisoformat(str(row[3])[:10])
            except (ValueError, TypeError):
                pass
        if len(row) > 4 and row[4]:
            try:
                valid_until = row[4] if isinstance(row[4], date) else date.fromisoformat(str(row[4])[:10])
            except (ValueError, TypeError):
                pass

        # 거래처 조회
        cust_result = await db.execute(
            select(Customer).where(Customer.code == customer_code, Customer.is_active.is_(True))
        )
        customer = cust_result.scalar_one_or_none()
        if not customer:
            errors.append(f"행 {row_idx}: 거래처코드 '{customer_code}' 없음")
            continue

        # 품목 조회
        prod_result = await db.execute(
            select(Product).where(Product.code == product_code, Product.is_active.is_(True))
        )
        product = prod_result.scalar_one_or_none()
        if not product:
            errors.append(f"행 {row_idx}: 품목코드 '{product_code}' 없음")
            continue

        # 기존 레코드 확인 (upsert)
        existing_result = await db.execute(
            select(CustomerPriceList).where(
                CustomerPriceList.customer_id == customer.id,
                CustomerPriceList.product_id == product.id,
                CustomerPriceList.is_active.is_(True),
            )
        )
        existing = existing_result.scalar_one_or_none()

        if existing:
            existing.unit_price = price
            existing.valid_from = valid_from
            existing.valid_until = valid_until
            updated += 1
        else:
            pl = CustomerPriceList(
                customer_id=customer.id,
                product_id=product.id,
                unit_price=price,
                valid_from=valid_from,
                valid_until=valid_until,
                created_by=current_user.id,
            )
            db.add(pl)
            created += 1

    await db.flush()
    wb.close()

    return {
        "message": f"거래처별 판매가 — 신규 {created}건, 수정 {updated}건 처리 완료",
        "created": created,
        "updated": updated,
        "errors": errors,
    }


# ── 엑셀 다운로드: 거래처별 판매가 템플릿 ──

async def download_template(db: AsyncSession, include_data: bool = False) -> bytes:
    """엑셀 템플릿 다운로드 (빈 양식 또는 현재 데이터 포함)"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()

    # 시트 1: 기본 판매가
    ws1 = wb.active
    ws1.title = "기본판매가"
    headers1 = ["품목코드", "품목명", "판매가"]
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E79", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    if include_data:
        products_result = await db.execute(
            select(Product).where(Product.is_active.is_(True)).order_by(Product.code)
        )
        for row_idx, prod in enumerate(products_result.scalars().all(), start=2):
            ws1.cell(row=row_idx, column=1, value=prod.code)
            ws1.cell(row=row_idx, column=2, value=prod.name)
            ws1.cell(row=row_idx, column=3, value=float(prod.standard_price) if prod.standard_price else 0)

    # 시트 2: 거래처별 판매가
    ws2 = wb.create_sheet("거래처별판매가")
    headers2 = ["거래처코드", "품목코드", "판매가", "유효시작일", "유효종료일"]
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4E79", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    if include_data:
        from sqlalchemy.orm import selectinload
        pl_result = await db.execute(
            select(CustomerPriceList)
            .options(selectinload(CustomerPriceList.customer), selectinload(CustomerPriceList.product))
            .where(CustomerPriceList.is_active.is_(True))
            .order_by(CustomerPriceList.customer_id, CustomerPriceList.product_id)
        )
        for row_idx, pl in enumerate(pl_result.scalars().all(), start=2):
            ws2.cell(row=row_idx, column=1, value=pl.customer.code if pl.customer else "")
            ws2.cell(row=row_idx, column=2, value=pl.product.code if pl.product else "")
            ws2.cell(row=row_idx, column=3, value=float(pl.unit_price))
            ws2.cell(row=row_idx, column=4, value=pl.valid_from.isoformat() if pl.valid_from else "")
            ws2.cell(row=row_idx, column=5, value=pl.valid_until.isoformat() if pl.valid_until else "")

    # 열 너비 조정
    for ws in [ws1, ws2]:
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col) + 2
            ws.column_dimensions[col[0].column_letter].width = min(max_len, 30)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
