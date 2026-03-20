"""
M2 영업/수주 — 견적서 서비스 (비즈니스 로직)
"""
import uuid
from datetime import date
from typing import Optional

from sqlalchemy import select, func, or_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from fastapi import HTTPException, status

from ..models import Quotation, QuotationLine
from ..schemas.quotations import QuotationCreate, QuotationUpdate
from ....audit.service import log_action
from ...m1_system.models import Customer


async def _generate_quote_no(db: AsyncSession, quote_date: date) -> str:
    """견적번호 자동 생성 (QT-YYYYMM-NNNN)"""
    prefix = f"QT-{quote_date.strftime('%Y%m')}-"
    result = await db.execute(
        select(func.count())
        .select_from(Quotation)
        .where(Quotation.quote_no.like(f"{prefix}%"))
    )
    seq = (result.scalar() or 0) + 1
    return f"{prefix}{seq:04d}"


def _calc_line(line_data) -> tuple[float, float]:
    """라인 금액 계산 → (공급가, 부가세)"""
    amount = round(line_data.quantity * line_data.unit_price * (1 - line_data.discount_rate / 100), 2)
    tax = round(amount * 0.1, 2)
    return amount, tax


async def list_quotations(
    db: AsyncSession,
    customer_id: Optional[str] = None,
    status_filter: Optional[str] = None,
    search: Optional[str] = None,
    page: int = 1,
    size: int = 20,
):
    """견적서 목록 조회"""
    query = select(Quotation).where(Quotation.is_active.is_(True))

    if customer_id:
        query = query.where(Quotation.customer_id == uuid.UUID(customer_id))
    if status_filter:
        query = query.where(Quotation.status == status_filter)
    if search:
        sf = f"%{search}%"
        query = query.join(Customer, Quotation.customer_id == Customer.id, isouter=True)
        query = query.where(or_(
            Quotation.quote_no.ilike(sf),
            Quotation.notes.ilike(sf),
            Customer.name.ilike(sf),
        ))

    # 전체 건수
    count_q = select(func.count()).select_from(query.subquery())
    total = (await db.execute(count_q)).scalar() or 0

    # 정렬 + 페이지네이션
    query = (
        query
        .options(selectinload(Quotation.customer), selectinload(Quotation.sales_rep))
        .order_by(Quotation.created_at.desc())
        .offset((page - 1) * size)
        .limit(size)
    )
    items = (await db.execute(query)).scalars().all()

    return {
        "items": [
            {
                "id": str(q.id),
                "quote_no": q.quote_no,
                "quote_date": q.quote_date,
                "valid_until": q.valid_until,
                "customer_name": q.customer.name if q.customer else None,
                "sales_rep_name": q.sales_rep.name if q.sales_rep else None,
                "total_amount": float(q.total_amount or 0),
                "tax_amount": float(q.tax_amount or 0),
                "grand_total": float(q.grand_total or 0),
                "status": q.status,
                "created_at": q.created_at,
            }
            for q in items
        ],
        "total": total,
        "page": page,
        "size": size,
        "total_pages": (total + size - 1) // size if total > 0 else 1,
    }


async def get_quotation(db: AsyncSession, quotation_id: uuid.UUID):
    """견적서 상세 조회 (라인 포함)"""
    result = await db.execute(
        select(Quotation)
        .options(
            selectinload(Quotation.lines).selectinload(QuotationLine.product),
            selectinload(Quotation.customer),
            selectinload(Quotation.sales_rep),
        )
        .where(Quotation.id == quotation_id, Quotation.is_active.is_(True))
    )
    q = result.scalar_one_or_none()
    if not q:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "견적서를 찾을 수 없습니다")

    return {
        "id": str(q.id),
        "quote_no": q.quote_no,
        "quote_date": q.quote_date,
        "valid_until": q.valid_until,
        "customer_id": str(q.customer_id),
        "customer_name": q.customer.name if q.customer else None,
        "sales_rep_id": str(q.sales_rep_id) if q.sales_rep_id else None,
        "sales_rep_name": q.sales_rep.name if q.sales_rep else None,
        "total_amount": float(q.total_amount or 0),
        "tax_amount": float(q.tax_amount or 0),
        "grand_total": float(q.grand_total or 0),
        "status": q.status,
        "notes": q.notes,
        "lines": [
            {
                "id": str(ln.id),
                "line_no": ln.line_no,
                "product_id": str(ln.product_id) if ln.product_id else None,
                "product_name": ln.product_name,
                "specification": ln.specification,
                "quantity": float(ln.quantity),
                "unit_price": float(ln.unit_price),
                "discount_rate": float(ln.discount_rate or 0),
                "amount": float(ln.amount),
                "tax_amount": float(ln.tax_amount or 0),
                "delivery_date": ln.delivery_date,
                "remark": ln.remark,
            }
            for ln in q.lines
        ],
        "created_at": q.created_at,
    }


async def create_quotation(
    db: AsyncSession,
    data: QuotationCreate,
    current_user,
    ip_address: Optional[str] = None,
):
    """견적서 생성"""
    quote_no = await _generate_quote_no(db, data.quote_date)

    # 헤더 생성
    quotation = Quotation(
        quote_no=quote_no,
        quote_date=data.quote_date,
        valid_until=data.valid_until,
        customer_id=uuid.UUID(data.customer_id),
        sales_rep_id=uuid.UUID(data.sales_rep_id) if data.sales_rep_id else None,
        notes=data.notes,
        status="draft",
        created_by=current_user.id,
    )
    db.add(quotation)
    await db.flush()

    # 라인 생성 + 합계 계산
    total_amount = 0.0
    total_tax = 0.0
    for idx, line_data in enumerate(data.lines, 1):
        amount, tax = _calc_line(line_data)
        line = QuotationLine(
            quotation_id=quotation.id,
            line_no=idx,
            product_id=uuid.UUID(line_data.product_id) if line_data.product_id else None,
            product_name=line_data.product_name,
            specification=line_data.specification,
            quantity=line_data.quantity,
            unit_price=line_data.unit_price,
            discount_rate=line_data.discount_rate,
            amount=amount,
            tax_amount=tax,
            delivery_date=line_data.delivery_date,
            remark=line_data.remark,
        )
        db.add(line)
        total_amount += amount
        total_tax += tax

    quotation.total_amount = round(total_amount, 2)
    quotation.tax_amount = round(total_tax, 2)
    quotation.grand_total = round(total_amount + total_tax, 2)
    await db.flush()

    # 감사 로그
    await log_action(
        db=db, table_name="quotations", record_id=quotation.id,
        action="INSERT", changed_by=current_user.id,
        new_values={"quote_no": quote_no, "customer_id": data.customer_id},
        ip_address=ip_address,
    )

    return {"id": str(quotation.id), "quote_no": quote_no}


async def update_quotation(
    db: AsyncSession,
    quotation_id: uuid.UUID,
    data: QuotationUpdate,
    current_user,
    ip_address: Optional[str] = None,
):
    """견적서 수정 (draft 상태만)"""
    result = await db.execute(
        select(Quotation)
        .options(selectinload(Quotation.lines))
        .where(Quotation.id == quotation_id, Quotation.is_active.is_(True))
    )
    q = result.scalar_one_or_none()
    if not q:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "견적서를 찾을 수 없습니다")
    if q.status != "draft":
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "draft 상태에서만 수정할 수 있습니다")

    # 헤더 필드 업데이트
    if data.quote_date is not None:
        q.quote_date = data.quote_date
    if data.valid_until is not None:
        q.valid_until = data.valid_until
    if data.customer_id is not None:
        q.customer_id = uuid.UUID(data.customer_id)
    if data.sales_rep_id is not None:
        q.sales_rep_id = uuid.UUID(data.sales_rep_id) if data.sales_rep_id else None
    if data.notes is not None:
        q.notes = data.notes

    # 라인 교체
    if data.lines is not None:
        # 기존 라인 삭제
        for old_line in q.lines:
            await db.delete(old_line)
        await db.flush()

        total_amount = 0.0
        total_tax = 0.0
        for idx, line_data in enumerate(data.lines, 1):
            amount, tax = _calc_line(line_data)
            line = QuotationLine(
                quotation_id=q.id,
                line_no=idx,
                product_id=uuid.UUID(line_data.product_id) if line_data.product_id else None,
                product_name=line_data.product_name,
                specification=line_data.specification,
                quantity=line_data.quantity,
                unit_price=line_data.unit_price,
                discount_rate=line_data.discount_rate,
                amount=amount,
                tax_amount=tax,
                delivery_date=line_data.delivery_date,
                remark=line_data.remark,
            )
            db.add(line)
            total_amount += amount
            total_tax += tax

        q.total_amount = round(total_amount, 2)
        q.tax_amount = round(total_tax, 2)
        q.grand_total = round(total_amount + total_tax, 2)

    await db.flush()

    await log_action(
        db=db, table_name="quotations", record_id=q.id,
        action="UPDATE", changed_by=current_user.id,
        new_values={"quote_no": q.quote_no},
        ip_address=ip_address,
    )

    return {"id": str(q.id), "quote_no": q.quote_no}


async def delete_quotation(
    db: AsyncSession,
    quotation_id: uuid.UUID,
    current_user,
    ip_address: Optional[str] = None,
):
    """견적서 삭제 (draft 상태만)"""
    result = await db.execute(
        select(Quotation).where(
            Quotation.id == quotation_id, Quotation.is_active.is_(True)
        )
    )
    q = result.scalar_one_or_none()
    if not q:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "견적서를 찾을 수 없습니다")
    if q.status != "draft":
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "draft 상태에서만 삭제할 수 있습니다")

    q.is_active = False

    await log_action(
        db=db, table_name="quotations", record_id=q.id,
        action="DELETE", changed_by=current_user.id,
        old_values={"quote_no": q.quote_no},
        ip_address=ip_address,
    )

    return {"message": f"견적서 {q.quote_no}이(가) 삭제되었습니다"}


async def update_quotation_status(
    db: AsyncSession,
    quotation_id: uuid.UUID,
    new_status: str,
    current_user,
    ip_address: Optional[str] = None,
):
    """견적서 상태 변경 (draft→sent→accepted/rejected)"""
    valid_transitions = {
        "draft": ["sent"],
        "sent": ["accepted", "rejected"],
    }

    result = await db.execute(
        select(Quotation).where(
            Quotation.id == quotation_id, Quotation.is_active.is_(True)
        )
    )
    q = result.scalar_one_or_none()
    if not q:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "견적서를 찾을 수 없습니다")

    allowed = valid_transitions.get(q.status, [])
    if new_status not in allowed:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            f"'{q.status}' → '{new_status}' 상태 전환이 불가합니다. 허용: {allowed}",
        )

    old_status = q.status
    q.status = new_status
    await db.flush()

    await log_action(
        db=db, table_name="quotations", record_id=q.id,
        action="UPDATE", changed_by=current_user.id,
        old_values={"status": old_status},
        new_values={"status": new_status},
        ip_address=ip_address,
        memo=f"상태 변경: {old_status} → {new_status}",
    )

    return {"id": str(q.id), "status": new_status}


async def generate_quotation_excel(db: AsyncSession, quotation_id: uuid.UUID) -> bytes:
    """견적서 Excel 양식 생성 (한국 표준 양식)"""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── 데이터 조회 ──
    result = await db.execute(
        select(Quotation)
        .options(
            selectinload(Quotation.lines).selectinload(QuotationLine.product),
            selectinload(Quotation.customer),
            selectinload(Quotation.sales_rep),
        )
        .where(Quotation.id == quotation_id, Quotation.is_active.is_(True))
    )
    q = result.scalar_one_or_none()
    if not q:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "견적서를 찾을 수 없습니다")

    # 회사 정보 조회
    from ...m1_system.models import CompanyInfo
    ci_result = await db.execute(select(CompanyInfo).limit(1))
    ci = ci_result.scalar_one_or_none()

    # ── 워크북 생성 ──
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "견적서"

    # 스타일 정의
    thin = Side(style="thin")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font = Font(name="맑은 고딕", size=20, bold=True)
    header_font = Font(name="맑은 고딕", size=10, bold=True)
    normal_font = Font(name="맑은 고딕", size=10)
    small_font = Font(name="맑은 고딕", size=9)
    center = Alignment(horizontal="center", vertical="center")
    left_al = Alignment(horizontal="left", vertical="center")
    right_al = Alignment(horizontal="right", vertical="center")
    header_fill = PatternFill(start_color="D9E2F3", fill_type="solid")

    # 열 너비 설정 (A~G: No, 품명, 규격, 수량, 단가, 금액, 비고)
    col_widths = [6, 28, 14, 10, 14, 16, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    row = 1

    # ── 제목 ──
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    cell = ws.cell(row=row, column=1, value="견   적   서")
    cell.font = title_font
    cell.alignment = center
    row += 2

    # ── 회사 정보 (좌측) + 수신처 (우측) ──
    company_name = ci.company_name if ci else "(주)피엘에스"
    business_no = ci.business_no if ci else ""
    ceo_name = ci.ceo_name if ci else ""
    address = ci.address if ci else ""
    biz_type = ci.business_type if ci else ""
    biz_item = ci.business_item if ci else ""
    phone = ci.phone if ci else ""
    fax = ci.fax if ci else ""

    # 공급자 정보
    info_data = [
        ("공급자", company_name),
        ("사업자번호", business_no),
        ("대 표 자", ceo_name),
        ("주    소", address),
        ("업태 / 종목", f"{biz_type} / {biz_item}"),
        ("TEL / FAX", f"{phone or ''} / {fax or ''}"),
    ]
    for label, value in info_data:
        ws.cell(row=row, column=1, value=label).font = header_font
        ws.cell(row=row, column=1).alignment = right_al
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.cell(row=row, column=2, value=value).font = normal_font
        ws.cell(row=row, column=2).alignment = left_al
        row += 1

    row += 1

    # ── 수신처 + 견적 정보 ──
    customer_name = q.customer.name if q.customer else ""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row=row, column=1, value=f"수  신 : {customer_name} 귀하").font = Font(name="맑은 고딕", size=12, bold=True)
    row += 1

    qt_info = [
        ("견적번호", q.quote_no),
        ("견적일자", str(q.quote_date)),
        ("유효기간", str(q.valid_until) if q.valid_until else "-"),
        ("합계금액", f"₩{q.grand_total:,.0f} (부가세 포함)"),
    ]
    for label, value in qt_info:
        ws.cell(row=row, column=1, value=label).font = header_font
        ws.cell(row=row, column=1).alignment = right_al
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        ws.cell(row=row, column=2, value=value).font = normal_font
        row += 1

    row += 1

    # ── 품목 테이블 헤더 ──
    headers = ["No", "품명", "규격", "수량", "단가", "금액", "비고"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = header_font
        cell.alignment = center
        cell.fill = header_fill
        cell.border = border_all
    row += 1

    # ── 품목 라인 ──
    for ln in q.lines:
        ws.cell(row=row, column=1, value=ln.line_no).font = normal_font
        ws.cell(row=row, column=1).alignment = center
        ws.cell(row=row, column=2, value=ln.product_name).font = normal_font
        ws.cell(row=row, column=3, value=ln.specification or "").font = normal_font
        ws.cell(row=row, column=4, value=float(ln.quantity)).font = normal_font
        ws.cell(row=row, column=4).alignment = right_al
        ws.cell(row=row, column=4).number_format = '#,##0'
        ws.cell(row=row, column=5, value=float(ln.unit_price)).font = normal_font
        ws.cell(row=row, column=5).alignment = right_al
        ws.cell(row=row, column=5).number_format = '#,##0'
        ws.cell(row=row, column=6, value=float(ln.amount)).font = normal_font
        ws.cell(row=row, column=6).alignment = right_al
        ws.cell(row=row, column=6).number_format = '#,##0'
        ws.cell(row=row, column=7, value=ln.remark or "").font = small_font
        for col_idx in range(1, 8):
            ws.cell(row=row, column=col_idx).border = border_all
        row += 1

    # 빈 줄 채우기 (최소 10줄)
    lines_count = len(q.lines)
    for _ in range(max(0, 10 - lines_count)):
        for col_idx in range(1, 8):
            ws.cell(row=row, column=col_idx).border = border_all
        row += 1

    row += 1

    # ── 합계 영역 ──
    summary = [
        ("공급가액", float(q.total_amount or 0)),
        ("부 가 세", float(q.tax_amount or 0)),
        ("합    계", float(q.grand_total or 0)),
    ]
    for label, value in summary:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = header_font
        cell.alignment = right_al
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=7)
        val_cell = ws.cell(row=row, column=4, value=value)
        val_cell.font = Font(name="맑은 고딕", size=11, bold=True) if label == "합    계" else normal_font
        val_cell.alignment = right_al
        val_cell.number_format = '₩#,##0'
        for col_idx in range(1, 8):
            ws.cell(row=row, column=col_idx).border = border_all
        row += 1

    # ── 비고 ──
    if q.notes:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        ws.cell(row=row, column=1, value=f"비고: {q.notes}").font = normal_font

    # ── 인쇄 설정 ──
    ws.print_area = f"A1:G{row + 1}"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    # ── 저장 ──
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
